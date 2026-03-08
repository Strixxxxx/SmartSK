from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import base64
import os
import io
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import logging

logger = logging.getLogger(__name__)

def rgb_to_hex(color):
    """Converts openpyxl color to hex string."""
    if not color:
        return None
    try:
        if color.type == 'rgb':
            rgb = color.rgb
            if len(rgb) == 8: # AARRGGBB
                return "#" + rgb[2:]
            return "#" + rgb
    except:
        pass
    return None

def get_border_style(border_side):
    """Maps openpyxl border style to Luckysheet border type."""
    if not border_side or not border_side.style:
        return None
    style_map = {
        'hair': 1, 'dotted': 2, 'dashDotDot': 3, 'dashDot': 4,
        'dashed': 5, 'mediumDashDotDot': 6, 'slantDashDot': 7,
        'mediumDashDot': 8, 'mediumDashed': 9, 'thin': 1,
        'medium': 2, 'thick': 3, 'double': 4
    }
    ls_style = style_map.get(border_side.style, 1)
    color = rgb_to_hex(border_side.color) or "#000000"
    return {"style": ls_style, "color": color}

def excel_to_fortune_json(file_input):
    """
    Converts an XLSX file (path, bytes, or BytesIO) to FortuneSheet (Luckysheet) compatible JSON.
    Focuses on 1:1 fidelity for smartSK templates including logos and styles.
    """
    print(f"DEBUG: Executing excel_to_fortune_json from {__file__}")
    
    if isinstance(file_input, (bytes, io.BytesIO)):
        if isinstance(file_input, bytes):
            file_input = io.BytesIO(file_input)
        wb = load_workbook(file_input, data_only=True)
    else:
        if not os.path.exists(file_input):
            raise FileNotFoundError(f"File not found: {file_input}")
        wb = load_workbook(file_input, data_only=True)
    sheets_json = []

    for index, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        
        sheet_data = {
            "name": sheet_name,
            "color": "",
            "status": 1 if index == 0 else 0,
            "order": index,
            "celldata": [],
            "config": {
                "merge": {},
                "rowlen": {},
                "columnlen": {},
                "borderInfo": []
            },
            "images": []
        }

        # Cache row/col dimensions for image positioning
        row_heights = {}
        col_widths = {}

        # 1. Process Cells and Dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(1, max_row + 1):
            h = ws.row_dimensions[r].height or 15 # Default 15pt
            h_px = int(h * 1.33)
            row_heights[r] = h_px
            sheet_data["config"]["rowlen"][str(r-1)] = h_px

        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            # openpyxl width is approx chars. Heuristic: char * 8px
            w = ws.column_dimensions[col_letter].width or 8.43
            w_px = int(w * 8) 
            col_widths[c] = w_px
            sheet_data["config"]["columnlen"][str(c-1)] = w_px

        # Process Merged Cells upfront to map all children
        merged_cells_map = {}
        for merge_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            r_idx, c_idx = min_row - 1, min_col - 1
            rs, cs = max_row - min_row + 1, max_col - min_col + 1
            sheet_data["config"]["merge"][f"{r_idx}_{c_idx}"] = {
                "r": r_idx, "c": c_idx, "rs": rs, "cs": cs
            }
            # Add mapping for every cell in this merge
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    if rr == min_row and cc == min_col:
                        merged_cells_map[(rr, cc)] = {"r": r_idx, "c": c_idx, "rs": rs, "cs": cs}
                    else:
                        merged_cells_map[(rr, cc)] = {"r": r_idx, "c": c_idx}

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                
                v_obj = {}
                
                # Check for merge details
                if (r, c) in merged_cells_map:
                    v_obj["mc"] = merged_cells_map[(r, c)]
                    
                # Skip full processing if it's a merged child (but we MUST append it with its mc object)
                if isinstance(cell, MergedCell):
                    if (r, c) in merged_cells_map:
                        sheet_data["celldata"].append({"r": r-1, "c": c-1, "v": v_obj})
                    continue
                
                # Value
                if cell.value is not None:
                    v_obj["v"] = cell.value
                    v_obj["m"] = str(cell.value)
                
                # Styles
                if cell.has_style:
                    # Font attributes
                    f = cell.font
                    if f.name: v_obj["ff"] = f.name
                    if f.sz: v_obj["fs"] = int(f.sz)
                    if f.b: v_obj["bl"] = 1
                    if f.i: v_obj["it"] = 1
                    if f.u: v_obj["un"] = 1
                    fc = rgb_to_hex(f.color)
                    if fc: v_obj["fc"] = fc

                    # Fill (Background)
                    if cell.fill and cell.fill.start_color:
                        bg = rgb_to_hex(cell.fill.start_color)
                        if bg and bg != "#000000": v_obj["bg"] = bg

                    # Alignment
                    if cell.alignment:
                        a = cell.alignment
                        # horizontal: 0 center, 1 left, 2 right
                        h_map = {"center": 0, "left": 1, "right": 2}
                        if a.horizontal in h_map: v_obj["ht"] = h_map[a.horizontal]
                        # vertical: 0 middle, 1 top, 2 bottom
                        v_map = {"center": 0, "top": 1, "bottom": 2}
                        if a.vertical in v_map: v_obj["vt"] = v_map[a.vertical]
                        # Wrap text
                        if a.wrapText: v_obj["tb"] = 2

                if sheet_name == "Governance" and r == 8 and c == 2:
                    print("GOV 8,2 BORDER INFO:", repr(cell.border))

                # Borders
                if cell.border:
                    b_info = {"row_index": r-1, "col_index": c-1}
                    has_border = False
                    if cell.border.left and getattr(cell.border.left, "style", None):
                        print(f"CELL {r}, {c} has border left: {cell.border.left.style}")
                        has_border = True; b_info["l"] = get_border_style(cell.border.left)
                    if cell.border.right and cell.border.right.style:
                        has_border = True; b_info["r"] = get_border_style(cell.border.right)
                    if cell.border.top and cell.border.top.style:
                        has_border = True; b_info["t"] = get_border_style(cell.border.top)
                    if cell.border.bottom and cell.border.bottom.style:
                        has_border = True; b_info["b"] = get_border_style(cell.border.bottom)
                        
                    if has_border:
                        sheet_data["config"]["borderInfo"].append({
                            "rangeType": "cell",
                            "value": b_info
                        })

                if v_obj:
                    sheet_data["celldata"].append({"r": r-1, "c": c-1, "v": v_obj})

        # 2. Process Images (Logos)
        for i, img in enumerate(ws._images):
            try:
                img_io = io.BytesIO()
                # Use PIL to normalize image
                pillow_img = PILImage.open(img.ref)
                pillow_img.save(img_io, format='PNG')
                b64_str = base64.b64encode(img_io.getvalue()).decode('utf-8')

                # Calculate position based on anchor
                top_offset = 0
                left_offset = 0
                
                if hasattr(img.anchor, '_from'):
                    anchor = img.anchor._from
                    # Sum row heights up to anchor row
                    for r_idx in range(1, anchor.row + 1):
                        top_offset += row_heights.get(r_idx, 20)
                    # Sum col widths up to anchor col
                    for c_idx in range(1, anchor.col + 1):
                        left_offset += col_widths.get(c_idx, 64)
                    
                    # Add fine offsets (openpyxl offsets are in EMUs: 1px = 9525 EMUs)
                    top_offset += (anchor.rowOff // 9525)
                    left_offset += (anchor.colOff // 9525)

                # Calculate exact rendered width/height
                render_width = img.width
                render_height = img.height
                try:
                    if hasattr(img.anchor, 'ext') and img.anchor.ext:
                        render_width = int(img.anchor.ext.cx / 9525)
                        render_height = int(img.anchor.ext.cy / 9525)
                    elif hasattr(img.anchor, 'width') and img.anchor.width:
                        render_width = int(img.anchor.width)
                        render_height = int(img.anchor.height)
                except Exception as ex:
                    logger.warning(f"Failed to infer exact rendered image dimensions: {ex}")

                img_id = f"img_{index}_{i}"
                sheet_data["images"].append({
                    "id": img_id,
                    "type": "png",
                    "src": f"data:image/png;base64,{b64_str}",
                    "width": render_width,
                    "height": render_height,
                    "left": left_offset,
                    "top": top_offset,
                    "originWidth": render_width,
                    "originHeight": render_height,
                    "default": {
                        "width": render_width,
                        "height": render_height,
                        "left": left_offset,
                        "top": top_offset
                    }
                })
            except Exception as e:
                logger.error(f"Image extraction failed: {e}")

        print(f"Sheet '{sheet_name}' border count before append: {len(sheet_data['config']['borderInfo'])}")
        sheets_json.append(sheet_data)

    return sheets_json

import os
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
from PIL import Image as PILImage
import logging
import io

logger = logging.getLogger(__name__)

def resize_image_proportional(img_path, target_width_px):
    """
    Resizes an image proportionally given a target width in pixels.
    Returns an openpyxl Image object.
    """
    with PILImage.open(img_path) as img:
        # Calculate aspect ratio
        width, height = img.size
        aspect_ratio = height / width
        target_height_px = int(target_width_px * aspect_ratio)
        
        # We don't actually need to save a temporary file, openpyxl.drawing.image.Image 
        # can take a PIL Image object directly in some versions, but to be safe and 
        # consistent with openpyxl API, we'll set the width/height on the wrapper.
        ox_img = OpenpyxlImage(img_path)
        ox_img.width = target_width_px
        ox_img.height = target_height_px
        return ox_img

def duplicate_and_init_excel(file_data: bytes, barangay_id: int, proj_type: str, target_year: str, sk_logo_data: bytes, brgy_logo_data: bytes) -> bytes:
    """
    Updates target years across all sheets in the provided Excel file data,
    and programmatically re-inserts SK and Barangay logos with proportional resizing.
    Operates entirely in memory and returns the modified Excel data as bytes.
    """
    try:
        # Load from bytes
        wb = load_workbook(io.BytesIO(file_data))
        
        # 4. Iterate through all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # --- CELL UPDATES ---
            if proj_type == 'ABYIP':
                ws['B9'] = f"FY {target_year}"
            elif proj_type == 'CBYDP':
                ws['B10'] = f"COMPREHENSIVE BARANGAY YOUTH DEVELOPMENT PLAN (CBYDP) CY {target_year}"
                
                years = target_year.split('-')
                start_year = int(years[0]) if len(years) > 0 else 0
                if start_year > 0:
                    for i, col in enumerate(['E', 'F', 'G']):
                        cell = ws[f'{col}13']
                        cell.value = start_year + i
                        cell.alignment = Alignment(horizontal='center')

            # --- LOGO RE-INSERTION (RESIZED) ---
            # 1 pixel = 9525 EMUs (English Metric Units)
            pixel_to_emu = 9525
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            
            target_width = 86 if proj_type == 'CBYDP' else 100  # 0.9" vs 1.04"

            # Barangay Logo on D1
            if brgy_logo_data:
                img_brgy_io = io.BytesIO(brgy_logo_data)
                img_brgy = resize_image_proportional(img_brgy_io, target_width)
                # D is col index 3
                col_idx = 3 
                col_offset = 13 * pixel_to_emu if proj_type == 'CBYDP' else 0
                
                marker = AnchorMarker(col=col_idx, colOff=col_offset, row=0, rowOff=0)
                size = XDRPositiveSize2D(img_brgy.width * pixel_to_emu, img_brgy.height * pixel_to_emu)
                img_brgy.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img_brgy)
            
            # SK Logo on H1
            if sk_logo_data:
                img_sk_io = io.BytesIO(sk_logo_data)
                img_sk = resize_image_proportional(img_sk_io, target_width)
                # H is col index 7
                col_idx = 7
                col_offset = 13 * pixel_to_emu if proj_type == 'CBYDP' else 0
                
                marker = AnchorMarker(col=col_idx, colOff=col_offset, row=0, rowOff=0)
                size = XDRPositiveSize2D(img_sk.width * pixel_to_emu, img_sk.height * pixel_to_emu)
                img_sk.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img_sk)

        # 5. Save Final File to memory
        output = io.BytesIO()
        wb.save(output)
        logger.info(f"Successfully initialized Excel in memory.")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error in in-memory Excel duplication: {e}")
        return None

import os
import pyodbc
import openpyxl
from copy import copy
from datetime import datetime
import logging
from openpyxl.styles import Font, Border, Side, Alignment

from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

from database.db_utils import get_db_connection

from openpyxl.cell.cell import MergedCell

def normalize_center_name(name):
    """Maps long frontend/DB category names to the specific sheet names used in the Excel template."""
    mapping = {
        'General Administration Program': 'General Administration',
        'Maintenance and Other Operating Expenses': 'MOOE'
    }
    return mapping.get(name, name)

def calculate_row_height(text, font_size, column_width_chars):
    """
    Estimates height based on character count and column width.
    mimics the 'auto-expanding' behavior of the web UI.
    """
    if not text or str(text).strip() in ["", "None", "N/A"]:
        return 15 # Default base height
    
    import math
    text_str = str(text)
    # Average character width for Calibri/Standard fonts
    chars_per_line = max(1, column_width_chars * 1.1)
    
    # Calculate lines from wrapped text
    wrapped_lines = 0
    for block in text_str.split('\n'):
        wrapped_lines += max(1, math.ceil(len(block) / chars_per_line))
    
    # Base height per line
    line_height = font_size * 1.5
    total_height = wrapped_lines * line_height
    
    return max(15, total_height + 4)

def safe_write(ws, row, col, val, font=None, alignment=None, border=None):
    """Writes to a cell only if it is not a MergedCell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = val
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def copy_row_style(ws, source_row, target_row):
    """Copies row dimensions and cell styles from source_row to target_row."""
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for c in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=c)
        target_cell = ws.cell(row=target_row, column=c)
        
        if isinstance(target_cell, MergedCell):
            continue
            
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

def apply_all_borders(ws, start_row, end_row, start_col, end_col):
    """Applies a uniform thin border to every cell in the specified range."""
    thin = Side(border_style="thin", color="000000")
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            # Apply border regardless of merging to ensure visual consistency
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def update_cbydp_sheet(ws, db_rows, agenda_statement=""):
    """Updates a single CBYDP sheet dynamically."""
    
    # --- Write Agenda Statement ---
    agenda_cell = ws.cell(row=7, column=2)
    agenda_cell.value = str(agenda_statement) if agenda_statement else ""
    agenda_cell.font = Font(name='Calibri', size=11)
    agenda_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    # --- Write Rows ---
    section_anchors = {}
    # Scan from column B (2) for section headers
    for r in range(1, 100):
        val = str(ws.cell(row=r, column=2).value).strip()
        if val in ["FROM", "TO", "ADDITIONAL PROJECT"]:
            section_anchors[val] = r

    if not section_anchors:
        return

    data_by_section = {'FROM': [], 'TO': [], 'ADDITIONAL PROJECT': []}
    for row in db_rows:
        secType = row[11]
        if secType in data_by_section:
            data_by_section[secType].append(row)

    for sec in data_by_section:
        data_by_section[sec] = sorted(data_by_section[sec], key=lambda x: x[12])

    col_indices = range(2, 11) # B to J
    # Bottom to top processing
    sections_order = [
        ("ADDITIONAL PROJECT", "bottom"),
        ("TO", "ADDITIONAL PROJECT"),
        ("FROM", "TO")
    ]

    for sec_name, next_sec_name in sections_order:
        if sec_name not in section_anchors: continue
        header_row = section_anchors[sec_name]
        data_start_row = header_row + 1
        
        # Fix horizontal merge for section header
        try:
            for m_range in list(ws.merged_cells.ranges):
                if m_range.min_row == header_row and m_range.max_row == header_row:
                    ws.unmerge_cells(str(m_range))
            ws.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=10)
            cell = ws.cell(row=header_row, column=2)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            logger.warning(f"Could not format header merge for {sec_name}: {e}")

        # Calculate allocated rows in template
        if next_sec_name == "bottom":
            allocated_rows = 6
        else:
            allocated_rows = section_anchors[next_sec_name] - data_start_row

        target_rows = len(data_by_section[sec_name])
        rows_needed = max(1, target_rows)
        
        # Aggressively unmerge Column B in the affected data range
        try:
            # Look ahead for any merges in Column B up to a reasonable limit
            for m_range in list(ws.merged_cells.ranges):
                if m_range.min_col <= 2 and m_range.max_col >= 2:
                    # If the merge overlaps our current data section
                    if not (m_range.max_row < header_row or m_range.min_row > (data_start_row + rows_needed + 10)):
                        # If it's a vertical merge in column B, unmerge it
                        if m_range.min_col == 2 and m_range.max_col == 2:
                            ws.unmerge_cells(str(m_range))
        except Exception as e:
            logger.debug(f"Unmerge check: {e}")

        diff = rows_needed - allocated_rows

        if diff > 0:
            ws.insert_rows(data_start_row + 1, amount=diff)
            for nr in range(data_start_row + 1, data_start_row + 1 + diff):
                copy_row_style(ws, data_start_row, nr)
            for k in section_anchors:
                if section_anchors[k] > header_row: section_anchors[k] += diff
        elif diff < 0:
            ws.delete_rows(data_start_row, amount=abs(diff))
            for k in section_anchors:
                if section_anchors[k] > header_row: section_anchors[k] -= abs(diff)

        # Write
        db_data = data_by_section[sec_name]
        for r_idx in range(rows_needed):
            curr_row = data_start_row + r_idx
            max_h = 15
            for c_idx in col_indices:
                cell = ws.cell(row=curr_row, column=c_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = None
            
            if r_idx < len(db_data):
                vals = db_data[r_idx][1:10]
                for v_idx, val in enumerate(vals):
                    col_idx = col_indices[v_idx]
                    
                    display_val = "N/A" if val is None or str(val).strip() == "" else val

                    forced_bold = False
                    if col_idx == 9 and "MOOE" in str(display_val):
                        forced_bold = True

                    # User specified CBYDP width is 15
                    max_h = max(max_h, calculate_row_height(display_val, 8, 15))

                    # Budget column (col_idx 9) should be center-aligned
                    h_align = 'center' if col_idx == 9 else None
                    
                    safe_write(ws, curr_row, col_idx, display_val, 
                              font=Font(name='Calibri', size=8, bold=forced_bold),
                              alignment=Alignment(wrap_text=True, vertical='center', horizontal=h_align))
            else:
                for c_idx in col_indices:
                    safe_write(ws, curr_row, c_idx, "N/A", 
                              font=Font(name='Calibri', size=8),
                              alignment=Alignment(wrap_text=True, vertical='center'))
            
            ws.row_dimensions[curr_row].height = max_h

    # Set consistent column width of 15 for CBYDP
    for c_idx in range(2, 11):
        ws.column_dimensions[get_column_letter(c_idx)].width = 15

    if "FROM" in section_anchors:
        ap_header = section_anchors.get("ADDITIONAL PROJECT", 20)
        ap_rows = max(1, len(data_by_section["ADDITIONAL PROJECT"]))
        # Apply borders from Row 10 (Main CBYDP Header) to the end of data
        apply_all_borders(ws, 10, ap_header + ap_rows, 2, 10)

def update_abyip_sheet(ws, db_rows):
    """Updates a single ABYIP sheet dynamically."""
    # Data starts at row 14, Columns B (2) to L (12)
    start_row = 14 
    col_indices = range(2, 13)
    db_rows = sorted(db_rows, key=lambda x: x[13])
    
    # Template has 8 empty rows initially
    allocated_rows = 8 
    rows_needed = max(1, len(db_rows))
    diff = rows_needed - allocated_rows

    if diff > 0:
        ws.insert_rows(start_row + 1, amount=diff)
        for nr in range(start_row + 1, start_row + 1 + diff):
            copy_row_style(ws, start_row, nr)
    elif diff < 0:
        ws.delete_rows(start_row, amount=abs(diff))

    # Set ABYIP Column Widths
    # B-G (2-7) and K-L (11-12) -> 14
    # H-J (8-10) -> 7
    for c_idx in range(2, 8): ws.column_dimensions[get_column_letter(c_idx)].width = 14
    for c_idx in range(8, 11): ws.column_dimensions[get_column_letter(c_idx)].width = 7
    for c_idx in range(11, 13): ws.column_dimensions[get_column_letter(c_idx)].width = 14

    for r_idx in range(rows_needed):
        curr_row = start_row + r_idx
        max_h = 15
        for ci in col_indices:
            cell = ws.cell(row=curr_row, column=ci)
            if not isinstance(cell, MergedCell):
                cell.value = None
        
        if r_idx < len(db_rows):
            vals = db_rows[r_idx][1:12]
            for v_idx, val in enumerate(vals):
                col_idx = col_indices[v_idx]
                
                display_val = "N/A" if val is None or str(val).strip() == "" else val
                
                # Determine width for height calculation
                w_est = 14 if col_idx not in [8, 9, 10] else 7
                max_h = max(max_h, calculate_row_height(display_val, 8, w_est))
                
                safe_write(ws, curr_row, col_idx, display_val,
                          font=Font(name='Calibri', size=8),
                          alignment=Alignment(wrap_text=True, vertical='center'))
        else:
             for c_idx in col_indices:
                 safe_write(ws, curr_row, c_idx, "N/A", 
                           font=Font(name='Calibri', size=8),
                           alignment=Alignment(wrap_text=True, vertical='center'))
        
        ws.row_dimensions[curr_row].height = max_h

    # ABYIP headers start at row 8 (ANNUAL BARANGAY YOUTH INVESTMENT PROGRAM)
    apply_all_borders(ws, 8, start_row + rows_needed - 1, 2, 12)

def sync_excel_from_db(batch_id, file_path):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT projType FROM projectBatch WHERE batchID = ?", batch_id)
        batch = cursor.fetchone()
        if not batch: return False
        proj_type = batch[0]
        
        wb = openpyxl.load_workbook(file_path)
        
        if proj_type == 'CBYDP':
            # Fetch CBYDP rows
            cursor.execute("SELECT cbydpID, YDC, objective, performanceIndicator, target1, target2, target3, PPAs, budget, personResponsible, centerOfParticipation, sectionType, sheetRowIndex FROM projectCBYDP WHERE projbatchID = ?", batch_id)
            rows = cursor.fetchall()
            
            # Fetch projectAgenda
            cursor.execute("SELECT governance, active_citizenship, economic_empowerment, global_mobility, agriculture, environment, PBS, SIE, education, health, GAP, MOOE FROM projectAgenda WHERE batchID = ?", batch_id)
            agenda = cursor.fetchone()
            
            agenda_map = {}
            if agenda:
                agenda_map = {
                    'Governance': agenda[0] or "",
                    'Active Citizenship': agenda[1] or "",
                    'Economic Empowerment': agenda[2] or "",
                    'Global Mobility': agenda[3] or "",
                    'Agriculture': agenda[4] or "",
                    'Environment': agenda[5] or "",
                    'Peace Building and Security': agenda[6] or "",
                    'Social Inclusion and Equity': agenda[7] or "",
                    'Education': agenda[8] or "",
                    'Health': agenda[9] or "",
                    'General Administration Program': agenda[10] or "",
                    'Maintenance and Other Operating Expenses': agenda[11] or ""
                }

            rows_by_sheet = {}
            for r in rows:
                c = normalize_center_name(r[10]) # cbydpID, ..., centerOfParticipation is index 10
                if c not in rows_by_sheet: rows_by_sheet[c] = []
                rows_by_sheet[c].append(r)
            
            # Map sheet names back to the agenda keys if needed
            sheet_to_agenda_key = {
                'General Administration': 'General Administration Program',
                'MOOE': 'Maintenance and Other Operating Expenses'
            }

            for sn in wb.sheetnames:
                # Find the data using the long name if it exists, otherwise use sheet name directly
                agenda_key = sheet_to_agenda_key.get(sn, sn)
                update_cbydp_sheet(wb[sn], rows_by_sheet.get(sn, []), agenda_map.get(agenda_key, ""))

        elif proj_type == 'ABYIP':
            cursor.execute("SELECT abyipID, referenceCode, PPA, [Description], expectedResult, performanceIndicator, period, PS, MOOE, CO, total, personResponsible, centerOfParticipation, sheetRowIndex FROM projectABYIP WHERE projbatchID = ?", batch_id)
            rows = cursor.fetchall()
            rows_by_sheet = {}
            for r in rows:
                c = normalize_center_name(r[12]) # abyipID, ..., centerOfParticipation is index 12
                if c not in rows_by_sheet: rows_by_sheet[c] = []
                rows_by_sheet[c].append(r)
            for sn in wb.sheetnames: update_abyip_sheet(wb[sn], rows_by_sheet.get(sn, []))

        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Sync error: {e}")
        return False
    finally:
        conn.close()

def sync_all_active_projects():
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT batchID FROM projectBatch WHERE isArchived = 0")
        batches = cursor.fetchall()
        
        results = []
        for (b_id,) in batches:
            pass
        return results
    finally:
        conn.close()
